import copy
from datetime import datetime
import openpyxl

old_data = []  # чтобы сравнить дату с датой при запросе на заявку, чтобы сравнить и создать лист если даты разные


def greeting():
    global gauth, old_data
    cur_date = datetime.now().strftime("%d_%m_%Y")  #название листа

    data_month = copy.copy(cur_date.split("_")[1])  # копия для сравнений
    old_data.clear()  # чистим предыдущую дату. Список создали всесто обычной переменной, чтобы можно было импортировать без потеря данных
    old_data.append(data_month)
    try:
        # если файл есть дописываем
        book = openpyxl.load_workbook("my_book.xlsx")
    except:
        # если нет создаем
        book = openpyxl.Workbook()

        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in book.sheetnames:
            sheet = book.get_sheet_by_name(sheet_name)
            book.remove_sheet(sheet)
            numbr = 1

    # создаем лист
    sheet = book.create_sheet(f'{cur_date}', 0)

    sheet['O1'].value = '.'  # чтобы лист создалось ставим точку в ячейкуу О
    print("создан новый лист")

    book.save("my_book.xlsx")
    book.close()


def main():
    greeting()


if __name__ == '__main__':
    main()